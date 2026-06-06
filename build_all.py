"""Batch-build PageIndex trees for every PDF in a directory.

Usage
-----
    # build every PDF under FinanceBench/, write trees/<name>.tree.json
    # already-built trees are skipped automatically; pass --force to rebuild.
    python build_all.py FinanceBench --out-dir trees

    # run 8 PDFs in parallel (default is 4)
    python build_all.py FinanceBench --workers 8

    # only build PDFs that have an embedded TOC (no LLM fallback)
    python build_all.py FinanceBench --skip-on-no-toc

    # smoke-test on the first PDF before committing to the full run
    python build_all.py FinanceBench --limit 1
"""
import argparse
import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import Lock
from typing import Any, Dict

from openpyxl import Workbook, load_workbook

from pageindex import Config, PageIndex


def _is_successful_tree(path: Path) -> bool:
    """A tree counts as successfully built if it exists and parses as JSON."""
    if not path.exists() or path.stat().st_size == 0:
        return False
    try:
        with open(path, "r", encoding="utf-8") as f:
            json.load(f)
    except (OSError, json.JSONDecodeError):
        return False
    return True


_ACCURACY_HEADER = [
    "pdf_name", "mode_used", "verifiable_leaves",
    "initial_accuracy", "final_accuracy", "iterations", "status", "saved",
]


def _append_accuracy_row(xlsx_path: Path, row: list) -> None:
    """Append one row to the accuracy log; create the workbook (with header) if missing."""
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "tree_accuracy"
        ws.append(_ACCURACY_HEADER)
    ws.append(row)
    wb.save(xlsx_path)


def _build_one(pdf_path: Path, out_path: Path, config: Config,
               effective_mode: str, verbose: bool) -> Dict[str, Any]:
    """Worker: build the tree for one PDF. No file I/O on shared resources.

    Returns a dict describing the outcome:
      {status: "OK", saved: bool, metrics, mode, timings, duration, root_pages}
      {status: "FAIL", error: str}
    The caller (main thread) handles xlsx append, log write, and final
    success/failure bookkeeping so those shared writes don't race.
    """
    t0 = time.time()
    index = None
    try:
        index = PageIndex(config=config)
        index.build_structure(str(pdf_path), mode=effective_mode, verbose=verbose)
        metrics = index.metrics
        saved = metrics["final_accuracy"] >= 0.60
        if saved:
            index.split_oversized(verbose=verbose)
            index.summarize(verbose=verbose)
            index.save(str(out_path))
        else:
            index.close()
        return {
            "status": "OK", "saved": saved, "metrics": metrics,
            "mode": index.mode, "timings": index.timings or {},
            "duration": time.time() - t0,
            "root_pages": index.root.end_page if index.root else 0,
        }
    except Exception as err:  # noqa: BLE001 - keep batch alive across failures
        if index is not None:
            try:
                index.close()
            except Exception:  # noqa: BLE001 - already failing
                pass
        return {"status": "FAIL", "error": str(err),
                "duration": time.time() - t0}


def main() -> int:
    parser = argparse.ArgumentParser(description="Batch-build PageIndex trees.")
    parser.add_argument("input_dir", help="Directory containing PDFs (searched recursively).")
    parser.add_argument("--out-dir", default="trees",
                        help="Directory to write *.tree.json files into.")
    parser.add_argument("--mode",
                        choices=["auto", "toc", "embedded", "printed", "window"],
                        default="auto", help="How to build each tree (default: auto).")
    parser.add_argument("--force", action="store_true",
                        help="Rebuild even when a successful tree.json already exists.")
    parser.add_argument("--skip-on-no-toc", action="store_true",
                        help="Use only the embedded TOC; skip PDFs that lack one.")
    parser.add_argument("--limit", type=int,
                        help="Process at most N PDFs (useful for a smoke test).")
    parser.add_argument("--workers", type=int, default=4,
                        help="Number of PDFs to build concurrently (default: 4). "
                             "Each worker runs its own PageIndex/LLMClient; "
                             "tune to your provider rate limits.")
    parser.add_argument("--verbose", action="store_true",
                        help="Print per-node progress while building. With "
                             "concurrent workers the output from different "
                             "PDFs may interleave.")
    args = parser.parse_args()

    in_dir = Path(args.input_dir)
    if not in_dir.is_dir():
        print(f"Error: {in_dir} is not a directory.", file=sys.stderr)
        return 1

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    pdfs = sorted(in_dir.rglob("*.pdf"))
    if args.limit:
        pdfs = pdfs[: args.limit]
    print(f"Found {len(pdfs)} PDF(s) under {in_dir}")
    if not pdfs:
        return 0

    config = Config()
    effective_mode = "toc" if args.skip_on_no_toc else args.mode

    successes: list[str] = []
    skipped: list[str] = []
    failures: list[tuple[str, str]] = []
    rejected: list[tuple[str, str]] = []
    log_path = out_dir / "_build_log.txt"
    accuracy_path = out_dir / "tree_accuracy.xlsx"

    # Decide upfront which PDFs need building so the skip count is deterministic
    # and the workers only see fresh work.
    to_build: list[Path] = []
    for pdf_path in pdfs:
        out_path = out_dir / (pdf_path.stem + ".tree.json")
        if not args.force and _is_successful_tree(out_path):
            skipped.append(pdf_path.name)
            print(f"[skip] {pdf_path.name}: already built.")
        else:
            to_build.append(pdf_path)

    if not to_build:
        print(f"\nNothing to build. {len(skipped)} skipped.")
        return 0

    # Single lock guards every write to log + accuracy xlsx from worker
    # callbacks (main thread runs them, but they share the file).
    io_lock = Lock()

    with open(log_path, "a", encoding="utf-8") as log:
        with io_lock:
            log.write(f"\n--- batch run {time.strftime('%Y-%m-%d %H:%M:%S')} "
                      f"(workers={args.workers}) ---\n")

        print(f"\nBuilding {len(to_build)} PDF(s) with {args.workers} worker(s)...")
        n_to_build = len(to_build)

        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            future_to_pdf = {}
            for pdf_path in to_build:
                out_path = out_dir / (pdf_path.stem + ".tree.json")
                fut = executor.submit(
                    _build_one, pdf_path, out_path, config,
                    effective_mode, args.verbose,
                )
                future_to_pdf[fut] = (pdf_path, out_path)

            for i, fut in enumerate(as_completed(future_to_pdf), 1):
                pdf_path, out_path = future_to_pdf[fut]
                tag = f"[{i}/{n_to_build}] {pdf_path.name}"
                result = fut.result()

                if result["status"] == "FAIL":
                    line = f"FAIL {pdf_path.name}  error={result['error']}"
                    with io_lock:
                        print(line, file=sys.stderr)
                        log.write(line + "\n")
                        log.flush()
                    failures.append((pdf_path.name, result["error"]))
                    continue

                metrics = result["metrics"]
                saved = result["saved"]
                tm = result["timings"]
                acc = metrics["final_accuracy"]
                outcome = (f"-> {out_path.name}" if saved
                           else f"NOT SAVED (status={metrics['status']})")
                splits = (f" [build={tm.get('build', 0):.1f}s "
                          f"verify={tm.get('verify', 0):.1f}s "
                          f"split={tm.get('split', 0):.1f}s "
                          f"summarize={tm.get('summarize', 0):.1f}s]")
                line = (f"OK   {pdf_path.name}  mode={result['mode']}  "
                        f"pages={result['root_pages']}  acc={acc:.0%}  "
                        f"time={result['duration']:.1f}s{splits}  {outcome}")

                with io_lock:
                    print(f"{tag}: {line}")
                    log.write(line + "\n")
                    log.flush()
                    _append_accuracy_row(accuracy_path, [
                        pdf_path.name, result["mode"], metrics["verifiable"],
                        round(metrics["initial_accuracy"], 4),
                        round(metrics["final_accuracy"], 4),
                        metrics["iterations"], metrics["status"], saved,
                    ])

                if saved:
                    successes.append(pdf_path.name)
                else:
                    rejected.append((pdf_path.name, metrics["status"]))

    print(f"\nDone. {len(successes)} built, {len(skipped)} skipped, "
          f"{len(rejected)} rejected, {len(failures)} failed.")
    print(f"Log: {log_path}")
    print(f"Accuracy: {accuracy_path}")
    if rejected:
        print("\nRejected (accuracy below threshold):")
        for name, status in rejected:
            print(f"  - {name}: {status}")
    if failures:
        print("\nFailures:")
        for name, err in failures:
            print(f"  - {name}: {err}")
    return 0 if not failures else 2


if __name__ == "__main__":
    sys.exit(main())