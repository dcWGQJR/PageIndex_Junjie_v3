"""Re-run only the verdict=F rows in an existing online.py results xlsx.

Reads the same xlsx that `online.py` writes (rows are: original QA columns +
predicted_answer + retrieval_path + verdict), finds rows whose verdict is F,
re-runs the answer + judge pipeline against the (already-built) tree, and
updates those three cells in place. Other rows are left untouched.

Usage
-----
    python retry_failed.py
    python retry_failed.py --out my_results.xlsx --limit 5
"""
import argparse
import sys
import time
from dataclasses import replace
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from pageindex import Config, LLMClient
from pageindex.tree import Node, load_tree

from online import JUDGE_SYS, _judge_prompt, answer_multi


def _run_question(doc_name: str, question: str, expected_str: str,
                  args, config: Config, llm: LLMClient, answer_llm: LLMClient,
                  trees_dir: Path, tree_cache: Dict[str, Node]) -> Dict[str, Any]:
    """Answer one question and judge it.

    Returns {status, predicted, breadcrumb, verdict, elapsed} where status is
    "OK", "SKIP" (missing tree / blank fields) or "ERROR" (answer crashed).
    """
    if not doc_name or not question or not expected_str:
        return {"status": "SKIP", "predicted": "", "breadcrumb": "",
                "verdict": "SKIP", "elapsed": 0.0}

    tree_path = trees_dir / f"{doc_name}.tree.json"
    if not tree_path.exists():
        return {"status": "SKIP", "predicted": "", "breadcrumb": "",
                "verdict": "SKIP", "elapsed": 0.0,
                "reason": f"no tree at {tree_path}"}

    if doc_name not in tree_cache:
        tree_cache[doc_name] = load_tree(str(tree_path))
    root = tree_cache[doc_name]

    t0 = time.time()
    try:
        result = answer_multi(
            root, question, config, llm,
            beam_size=args.beam_size, max_leaves=args.max_leaves,
            verbose=args.verbose, answer_llm=answer_llm,
        )
    except Exception as err:  # noqa: BLE001 - caller decides what to log
        return {"status": "ERROR", "predicted": "", "breadcrumb": "",
                "verdict": "ERROR", "elapsed": time.time() - t0,
                "reason": str(err)[:300]}

    predicted = result["answer"]
    breadcrumb = result["breadcrumb"]

    try:
        judged = answer_llm.complete_json(
            JUDGE_SYS, _judge_prompt(question, expected_str, predicted)
        )
        verdict = str(judged.get("verdict", "")).strip().upper()
        if verdict not in ("T", "F"):
            verdict = "?"
    except Exception:  # noqa: BLE001 - record judge failures
        verdict = "?"

    return {"status": "OK", "predicted": predicted, "breadcrumb": breadcrumb,
            "verdict": verdict, "elapsed": time.time() - t0}


def retry_failed(args, out_path: Path, trees_dir: Path, config: Config,
                 llm: LLMClient, answer_llm: LLMClient) -> int:
    """Re-run only verdict=F rows in an existing results xlsx; update in place."""
    if not out_path.is_file():
        print(f"Error: no results file at {out_path}", file=sys.stderr)
        return 1

    out_wb = load_workbook(out_path)
    out_ws = out_wb.active
    header = [c.value for c in out_ws[1]]
    for required in ("doc_name", "question", "answer",
                     "predicted_answer", "retrieval_path", "verdict"):
        if required not in header:
            print(f"Error: column {required!r} missing from {out_path}",
                  file=sys.stderr)
            return 1

    # 1-indexed column numbers for openpyxl cell access.
    doc_c = header.index("doc_name") + 1
    q_c = header.index("question") + 1
    exp_c = header.index("answer") + 1
    pred_c = header.index("predicted_answer") + 1
    path_c = header.index("retrieval_path") + 1
    verd_c = header.index("verdict") + 1
    just_c = header.index("justification") + 1 if "justification" in header else None
    evid_c = header.index("evidence") + 1 if "evidence" in header else None

    f_rows: List[int] = []
    for r in range(2, out_ws.max_row + 1):
        if str(out_ws.cell(row=r, column=verd_c).value or "").strip().upper() == "F":
            f_rows.append(r)

    if not f_rows:
        print(f"No verdict=F rows in {out_path}.")
        return 0

    if args.limit:
        f_rows = f_rows[: args.limit]

    n_total = len(f_rows)
    print(f"Retrying {n_total} F row(s) from {out_path}")

    n_t = n_f = n_unknown = n_errored = 0
    tree_cache: Dict[str, Node] = {}

    for i, r in enumerate(f_rows, start=1):
        doc_name = out_ws.cell(row=r, column=doc_c).value
        question = out_ws.cell(row=r, column=q_c).value
        expected = out_ws.cell(row=r, column=exp_c).value
        expected_str = "" if expected is None else str(expected).strip()

        res = _run_question(
            str(doc_name) if doc_name else "",
            str(question) if question else "",
            expected_str,
            args, config, llm, answer_llm, trees_dir, tree_cache,
        )

        if res["status"] == "SKIP":
            reason = res.get("reason", "blank fields")
            print(f"[{i}/{n_total}] SKIP  {doc_name}  ({reason})")
            continue
        if res["status"] == "ERROR":
            n_errored += 1
            print(f"[{i}/{n_total}] ERR   {doc_name}: {res.get('reason', '')}")
            out_ws.cell(row=r, column=verd_c).value = "ERROR"
            if i % args.save_every == 0:
                out_wb.save(out_path)
            continue

        verdict = res["verdict"]
        if verdict == "T":
            n_t += 1
        elif verdict == "F":
            n_f += 1
        else:
            n_unknown += 1

        out_ws.cell(row=r, column=pred_c).value = res["predicted"]
        out_ws.cell(row=r, column=path_c).value = res["breadcrumb"]
        out_ws.cell(row=r, column=verd_c).value = verdict

        q_text = str(question)
        q_preview = (q_text[:60] + "...") if len(q_text) > 60 else q_text
        judged_so_far = n_t + n_f
        running = f"running {n_t}/{judged_so_far}" if judged_so_far else "running -"

        if verdict == "F":
            out_wb.save(out_path)
            justification = (out_ws.cell(row=r, column=just_c).value
                             if just_c else "")
            evidence = out_ws.cell(row=r, column=evid_c).value if evid_c else ""
            for field, value in [
                ("question", question),
                ("answer", expected_str),
                ("justification", justification),
                ("evidence", evidence),
                ("predicted_answer", res["predicted"]),
                ("retrieval_path", res["breadcrumb"]),
                ("verdict", verdict),
            ]:
                print(f"\n--- {field} ---\n{value}")

        if i % args.save_every == 0:
            out_wb.save(out_path)

        print(f"[{i}/{n_total}] {verdict}  {doc_name}  ({res['elapsed']:.1f}s)  "
              f"[{running}]  q={q_preview!r}")

    out_wb.save(out_path)
    total_judged = n_t + n_f
    print(f"\n{'=' * 60}")
    print(f"Retry done.  T={n_t}  F={n_f}  ?={n_unknown}  errored={n_errored}")
    if total_judged:
        print(f"Recovered (T / (T+F)) = {n_t}/{total_judged} = {n_t / total_judged:.1%}")
    print(f"Results updated in {out_path}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Re-run only the verdict=F rows in an existing results xlsx."
    )
    parser.add_argument("--out", default="financebench_results.xlsx",
                        help="Existing results xlsx to update in place.")
    parser.add_argument("--trees-dir", default="trees")
    parser.add_argument("--limit", type=int,
                        help="Process at most N of the F rows.")
    parser.add_argument("--beam-size", type=int, default=100,
                        help="Max children to follow at each tree level (default: 100).")
    parser.add_argument("--max-leaves", type=int, default=100,
                        help="Max sections to feed into the answer (default: 100).")
    parser.add_argument("--save-every", type=int, default=5,
                        help="Persist the output xlsx every N rows.")
    parser.add_argument("--verbose", action="store_true",
                        help="Print every routing decision.")
    args = parser.parse_args()

    out_path = Path(args.out)
    trees_dir = Path(args.trees_dir)

    config = Config()
    llm = LLMClient(config)
    # Final answer uses a stronger model than routing/judge.
    answer_model = "gpt-4o-2024-11-20" if config.provider == "openai" else "claude-opus-4-7"
    answer_llm = LLMClient(replace(config, model=answer_model))

    return retry_failed(args, out_path, trees_dir, config, llm, answer_llm)


if __name__ == "__main__":
    sys.exit(main())