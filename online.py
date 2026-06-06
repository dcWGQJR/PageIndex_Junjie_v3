"""Answer FinanceBench questions against pre-built PageIndex trees and grade them.

Uses FLAT-SELECTION retrieval: every node in the tree (parents + leaves) is
listed in a single LLM prompt with its title, page range, and summary; the
selection agent returns the indices of the nodes most likely to answer the
query. The final answer is generated from the combined text of every selected
node, so a question whose answer lives in (for example) "Note 18. Business
Segments" is not lost when the answer also needs the "Consolidated Statement
of Cash Flows" - both excerpts reach the answer LLM in one shot.

For each row of `financebench_subset_QA.xlsx`:
- Load the pre-built tree at `trees/<doc_name>.tree.json`.
- Run one selection call over every node in the tree; collect picked nodes.
- Concatenate the selected sections' text into one excerpt block; answer.
- Use an LLM judge to compare the predicted answer to the benchmark answer.
- Append a row to an output xlsx with: <all original columns> + predicted_answer
  + retrieval_path + verdict (T/F/?/SKIP/ERROR).

Usage
-----
    python online.py
    python online.py --limit 5
    python online.py --max-picks 5 --out results_top5.xlsx
"""
import argparse
import sys
import time
from dataclasses import replace
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from online_prompts import (
    ANSWER_MULTI_SYS, JUDGE_SYS, SELECTION_SYS, VERIFY_ANSWER_SYS,
    answer_user_multi, excerpts_block, judge_prompt, selection_user,
    verify_answer_prompt,
)
from pageindex import Config, LLMClient
from pageindex.tree import Node, load_tree


# --------------------------------------------------------------------------
# Retrieval: flat selection over every node in the tree
# --------------------------------------------------------------------------
def _flatten_nodes(root: Node) -> List[Tuple[List[Node], Node]]:
    """Walk the tree in document order and return (path, node) for every
    non-root node. `path` is root -> ... -> node (inclusive of root and node)."""
    out: List[Tuple[List[Node], Node]] = []

    def walk(node: Node, path: List[Node]) -> None:
        for c in node.children:
            new_path = path + [c]
            out.append((new_path, c))
            walk(c, new_path)

    walk(root, [root])
    return out


def _format_nodes_block(entries: List[Tuple[List[Node], Node]]) -> str:
    """Render every (path, node) as an indented numbered block for the LLM.

    Indentation reflects depth so the LLM can read the parent/child structure
    visually. Each entry shows index, title, page range, and summary.
    """
    lines: List[str] = []
    for i, (path, node) in enumerate(entries):
        # depth = number of ancestors above root (root itself is depth 0).
        depth = len(path) - 2 if len(path) >= 2 else 0
        indent = "  " * max(0, depth)
        lines.append(
            f'{indent}[{i}] "{node.title}" (pages {node.start_page}-{node.end_page})'
        )
        if node.summary:
            lines.append(f"{indent}    Summary: {node.summary}")
    return "\n".join(lines)


def select_nodes(root: Node, query: str, llm: LLMClient, max_picks: int,
                 verbose: bool = False) -> List[List[Node]]:
    """Show every node (parents + leaves) to the LLM in one prompt and let it
    pick which ones to use. Returns the path root -> ... -> selected node for
    each selected node, in the LLM's order of relevance."""
    entries = _flatten_nodes(root)
    if not entries:
        return []
    nodes_block = _format_nodes_block(entries)
    result = llm.complete_json(
        SELECTION_SYS, selection_user(query, nodes_block, max_picks),
    )
    result = result if isinstance(result, dict) else {}
    raw = result.get("node_indices") or []
    if not isinstance(raw, list):
        raw = []

    picked: List[List[Node]] = []
    seen_ids: set = set()
    for item in raw:
        try:
            ix = int(item)
        except (TypeError, ValueError):
            continue
        if not (0 <= ix < len(entries)):
            continue
        path, node = entries[ix]
        if id(node) in seen_ids:
            continue
        seen_ids.add(id(node))
        picked.append(path)
        if len(picked) >= max_picks:
            break

    if verbose:
        reasoning = str(result.get("reasoning", "")).strip()
        titles = [p[-1].title for p in picked]
        print(f"[select] picked {len(picked)} node(s): {titles}")
        if reasoning:
            print(f"[select] reasoning: {reasoning}")

    return picked


def _subtree_text(node: Node) -> str:
    """Concatenate `node.text` plus every descendant's text in document order.

    Each node's stored `text` is its own non-overlapping slice (leaf.text is
    the full page range; parent.text is the preamble paragraphs not covered
    by any child). Walking the subtree pre-order and joining yields the full
    text of the section, with no duplication.
    """
    pieces: List[str] = []
    if node.text and node.text.strip():
        pieces.append(node.text)
    for c in node.children:
        sub = _subtree_text(c)
        if sub:
            pieces.append(sub)
    return "\n\n".join(pieces)


def _selected_excerpt(path: List[Node]) -> str:
    """Build the excerpt for a selected node: ancestor preambles + its subtree.

    Ancestor `node.text` (the preamble before any child) is prepended so the
    selected section is read with the framing the parent set up (scope,
    definitions, accounting basis...). The selected node's full subtree
    follows.
    """
    pieces: List[str] = []
    for ancestor in path[:-1]:
        if ancestor.text and ancestor.text.strip():
            pieces.append(f'[Preamble from "{ancestor.title}"]\n{ancestor.text}')
    target = path[-1]
    sub = _subtree_text(target)
    if sub:
        pieces.append(sub)
    return "\n\n".join(pieces)


# --------------------------------------------------------------------------
# Answer + verify
# --------------------------------------------------------------------------
# Input-token budget for the answer prompt. gpt-4o-2024-11-20 has a 128k
# total-context window; we reserve ~8k for the response and tokenizer slop.
# Claude has a 200k window but the same budget is conservative enough to
# work for it too, so we keep one number.
_ANSWER_INPUT_BUDGET = 120000

# Lazy-loaded tiktoken encoder for token estimation. Falls back to a
# character-based heuristic when tiktoken isn't installed.
_TOKEN_ENCODER: Any = None


def _estimate_tokens(text: str) -> int:
    """Estimate token count of `text`.

    Uses tiktoken's cl100k_base (GPT-4 / gpt-4o tokenizer) when available -
    it's also a decent rough proxy for Claude's tokenizer for English text.
    Falls back to `len(text) / 3` (conservative; dense numeric text averages
    closer to 3 chars/token than the usual ~4).
    """
    global _TOKEN_ENCODER
    if _TOKEN_ENCODER is None:
        try:
            import tiktoken
            _TOKEN_ENCODER = tiktoken.get_encoding("cl100k_base")
        except Exception:  # noqa: BLE001 - any failure -> use heuristic
            _TOKEN_ENCODER = "fallback"
    if _TOKEN_ENCODER == "fallback":
        return len(text) // 3
    return len(_TOKEN_ENCODER.encode(text))


def _trim_to_budget(query: str, paths: List[List[Node]], excerpts: List[str],
                    budget: int, verbose: bool = False
                    ) -> Tuple[List[List[Node]], List[str]]:
    """Drop trailing (least-relevant) picks until the answer prompt fits.

    `select_nodes` returns paths ordered by relevance (most-relevant first),
    so `paths[-1]` is the safest to drop. Estimates the rendered system +
    user token count via `_estimate_tokens` and pops one pick at a time
    until the total is within `budget` or only one pick is left.

    If a single pick is itself oversized, this returns it unchanged - the
    caller's LLM call will then error out and the row is recorded as ERROR.
    Truncating a single excerpt is not attempted here.
    """
    sys_tokens = _estimate_tokens(ANSWER_MULTI_SYS)
    while len(paths) > 1:
        user = answer_user_multi(query, paths, excerpts)
        total = sys_tokens + _estimate_tokens(user)
        if total <= budget:
            return paths, excerpts
        if verbose:
            dropped = paths[-1][-1].title
            print(f"[trim] estimated {total} tokens > budget {budget}; "
                  f"dropping least-relevant pick {dropped!r} "
                  f"({len(paths) - 1} remaining)")
        paths = paths[:-1]
        excerpts = excerpts[:-1]
    return paths, excerpts



def _reanswer_with_feedback(query: str, paths: List[List[Node]],
                            excerpts: List[str], prior_answer: str,
                            issues: List[str], config: Config,
                            answer_llm: LLMClient) -> str:
    """Re-run the answer model with verifier feedback appended to the user prompt."""
    base = answer_user_multi(query, paths, excerpts)
    issues_block = "\n".join(f"  - {s}" for s in issues if s)
    feedback = (
        "\n\nA verification step flagged the following issues with a previous "
        "attempt at this question. Address each before producing your final "
        "answer.\n\n"
        f"PREVIOUS ANSWER:\n{prior_answer}\n\n"
        f"VERIFIER ISSUES:\n{issues_block}\n"
    )
    return answer_llm.complete(
        ANSWER_MULTI_SYS,
        base + feedback,
        max_tokens=config.answer_max_tokens,
    )


def verify_answer(query: str, paths: List[List[Node]], excerpts: List[str],
                  predicted: str, judge_llm: LLMClient,
                  verbose: bool = False) -> Dict[str, Any]:
    """Run the verifier critic. Returns {ok, issues, corrected_final_answer}."""
    try:
        result = judge_llm.complete_json(
            VERIFY_ANSWER_SYS,
            verify_answer_prompt(query, excerpts_block(paths, excerpts), predicted),
            max_tokens=1000,
        )
    except Exception as err:  # noqa: BLE001 - degrade to "ok" rather than block answer
        if verbose:
            print(f"[verify_answer] critic error (treating as OK): {err}")
        return {"ok": True, "issues": [], "corrected_final_answer": ""}
    if not isinstance(result, dict):
        return {"ok": True, "issues": [], "corrected_final_answer": ""}
    ok = bool(result.get("ok", True))
    issues = result.get("issues") or []
    if not isinstance(issues, list):
        issues = []
    corrected = str(result.get("corrected_final_answer", "")).strip()
    return {"ok": ok, "issues": [str(s) for s in issues], "corrected_final_answer": corrected}


def answer_multi(root: Node, query: str, config: Config,
                 llm: LLMClient, max_picks: int = 100,
                 verbose: bool = False,
                 answer_llm: Optional[LLMClient] = None,
                 judge_llm: Optional[LLMClient] = None,
                 verify: bool = True) -> Dict[str, Any]:
    """Retrieve, answer, optionally verify-and-regenerate.

    `judge_llm` runs the verifier critic and (separately) the T/F judge.
    Falls back to `answer_llm` when not provided, so older callers keep
    working. When `verify=True` and the verifier flags issues, one re-answer
    pass is run with the verifier's issues appended to the user prompt.
    Verifier failures (network/parse errors) degrade to OK so a flaky critic
    never blocks an otherwise-fine answer.
    """
    paths = select_nodes(root, query, llm, max_picks=max_picks, verbose=verbose)
    excerpts = [_selected_excerpt(p) for p in paths]
    # Token-budget guard: drop least-relevant picks until the prompt fits.
    paths, excerpts = _trim_to_budget(query, paths, excerpts,
                                      _ANSWER_INPUT_BUDGET, verbose=verbose)
    breadcrumb = "  |  ".join(" > ".join(n.title for n in path) for path in paths)
    use_answer_llm = answer_llm or llm
    use_judge_llm = judge_llm or use_answer_llm
    answer = use_answer_llm.complete(
        ANSWER_MULTI_SYS,
        answer_user_multi(query, paths, excerpts),
        max_tokens=config.answer_max_tokens,
    ).strip()

    verification: Optional[Dict[str, Any]] = None
    if verify:
        verification = verify_answer(query, paths, excerpts, answer, use_judge_llm,
                                     verbose=verbose)
        if not verification["ok"] and verification["issues"]:
            if verbose:
                preview = "; ".join(verification["issues"])[:200]
                print(f"[verify_answer] flagged: {preview}")
            answer = _reanswer_with_feedback(
                query, paths, excerpts, answer, verification["issues"],
                config, use_answer_llm,
            ).strip()

    return {
        "answer": answer,
        "breadcrumb": breadcrumb,
        "paths": paths,
        "verification": verification,
    }


# --------------------------------------------------------------------------
# Main pipeline (CLI + per-row orchestration)
# --------------------------------------------------------------------------
def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Answer FinanceBench questions against pre-built PageIndex trees."
    )
    parser.add_argument("--qa-file", default="FinanceBench/financebench_subset_QA.xlsx")
    parser.add_argument("--trees-dir", default="trees")
    parser.add_argument("--out", default="financebench_results.xlsx")
    parser.add_argument("--limit", type=int, help="Process at most N rows.")
    parser.add_argument("--max-picks", type=int, default=10,
                        help="Max nodes the selection agent may pick (default: 10).")
    parser.add_argument("--save-every", type=int, default=5,
                        help="Persist the output xlsx every N rows.")
    parser.add_argument("--verbose", action="store_true",
                        help="Print every routing decision.")
    parser.add_argument("--dump-false", action="store_true",
                        help="Print the full diagnostic dump (question / expected / "
                             "justification / evidence / predicted / retrieval path) "
                             "for every row judged F. Off by default - the per-row "
                             "verdict line still prints, the xlsx still records F, "
                             "just no detailed dump.")
    return parser.parse_args()


def _load_qa(qa_path: Path):
    """Open the QA workbook and validate required columns.

    Returns `(ws, header, col)` where `col` maps each required name to its
    0-indexed column. Raises ValueError with a printable message on a missing
    file or missing columns.
    """
    if not qa_path.is_file():
        raise ValueError(f"QA file not found: {qa_path}")
    wb = load_workbook(qa_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    required = ["doc_name", "question", "answer"]
    missing = [c for c in required if c not in header]
    if missing:
        raise ValueError(f"required columns missing from QA file: {missing}")
    col = {c: header.index(c) for c in required}
    return ws, header, col


def _make_output_workbook(header: List[str]) -> Tuple[Workbook, Any]:
    """Create the results workbook with the QA header + three extra columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(header + ["predicted_answer", "retrieval_path", "verdict"])
    return wb, ws


def _make_llm_clients(config: Config) -> Tuple[LLMClient, LLMClient, LLMClient]:
    """Build the (selection, answer, judge) LLM client triple.

    - Selection:    gpt-4o-2024-11-20 on OpenAI; Sonnet on Anthropic.
    - Answer:       gpt-4o-2024-11-20 on OpenAI; Opus on Anthropic.
    - Judge/verify: gpt-4o-2024-11-20 on OpenAI; Sonnet on Anthropic.
    """
    routing_model = "gpt-4o-2024-11-20" if config.provider == "openai" else "claude-sonnet-4-6"
    routing_llm = LLMClient(replace(config, model=routing_model))
    answer_model = "gpt-4o-2024-11-20" if config.provider == "openai" else "claude-opus-4-7"
    answer_llm = LLMClient(replace(config, model=answer_model))
    judge_model = "gpt-4o-2024-11-20" if config.provider == "openai" else "claude-sonnet-4-6"
    judge_llm = LLMClient(replace(config, model=judge_model))
    return routing_llm, answer_llm, judge_llm


def _judge_one(judge_llm: LLMClient, question: str, expected: str,
               predicted: str) -> str:
    """Return T / F / ? for one row. Falls back to ? on judge LLM errors."""
    try:
        judged = judge_llm.complete_json(
            JUDGE_SYS, judge_prompt(question, expected, predicted)
        )
    except Exception:  # noqa: BLE001 - record judge failures as ?
        return "?"
    verdict = str(judged.get("verdict", "")).strip().upper()
    return verdict if verdict in ("T", "F") else "?"


def _process_row(r: int, n_total: int, row: tuple, col: Dict[str, int],
                 trees_dir: Path, tree_cache: Dict[str, Node],
                 config: Config, routing_llm: LLMClient, answer_llm: LLMClient,
                 judge_llm: LLMClient,
                 args: argparse.Namespace) -> Dict[str, Any]:
    """Answer + judge one QA row.

    Returns a dict with keys `verdict` (one of T/F/?/SKIP/ERROR),
    `predicted`, `breadcrumb`, `duration`. SKIP/ERROR rows have empty
    `predicted`/`breadcrumb`; OK rows always carry a real verdict.
    """
    doc_name = row[col["doc_name"]]
    question = row[col["question"]]
    expected = row[col["answer"]]
    expected_str = "" if expected is None else str(expected).strip()

    if not doc_name or not question or not expected_str:
        return {"verdict": "SKIP", "predicted": "", "breadcrumb": "", "duration": 0.0}

    tree_path = trees_dir / f"{doc_name}.tree.json"
    if not tree_path.exists():
        print(f"[{r-1}/{n_total}] SKIP  {doc_name}  (no tree at {tree_path})")
        return {"verdict": "SKIP", "predicted": "", "breadcrumb": "", "duration": 0.0}

    if doc_name not in tree_cache:
        tree_cache[doc_name] = load_tree(str(tree_path))
    root = tree_cache[doc_name]

    t0 = time.time()
    try:
        result = answer_multi(
            root, str(question), config, routing_llm,
            max_picks=args.max_picks, verbose=args.verbose,
            answer_llm=answer_llm, judge_llm=judge_llm,
        )
    except Exception as err:  # noqa: BLE001 - keep batch alive
        print(f"[{r-1}/{n_total}] ERR   {doc_name}: {str(err)[:300]}")
        return {"verdict": "ERROR", "predicted": "", "breadcrumb": "",
                "duration": time.time() - t0}

    predicted = result["answer"]
    breadcrumb = result["breadcrumb"]
    verdict = _judge_one(judge_llm, str(question), expected_str, predicted)
    return {"verdict": verdict, "predicted": predicted, "breadcrumb": breadcrumb,
            "duration": time.time() - t0}


def _print_progress(r: int, n_total: int, row: tuple, col: Dict[str, int],
                    result: Dict[str, Any], counters: Dict[str, int]) -> None:
    """Print the one-line per-row progress message for an OK row."""
    doc_name = row[col["doc_name"]]
    q_text = str(row[col["question"]])
    q_preview = (q_text[:60] + "...") if len(q_text) > 60 else q_text
    judged_so_far = counters["T"] + counters["F"]
    running = f"running {counters['T']}/{judged_so_far}" if judged_so_far else "running -"
    print(f"[{r-1}/{n_total}] {result['verdict']}  {doc_name}  "
          f"({result['duration']:.1f}s)  [{running}]  q={q_preview!r}")


def _print_f_dump(row: tuple, header: List[str], col: Dict[str, int],
                  result: Dict[str, Any]) -> None:
    """Print the full diagnostic dump for one F-verdict row."""
    question = row[col["question"]]
    expected = row[col["answer"]]
    expected_str = "" if expected is None else str(expected).strip()
    justification = row[header.index("justification")] if "justification" in header else ""
    evidence = row[header.index("evidence")] if "evidence" in header else ""
    print(f"\n{'=' * 60}")
    print("F verdict - dumping row:")
    print(f"{'=' * 60}")
    for field, value in [
        ("question", question),
        ("answer", expected_str),
        ("justification", justification),
        ("evidence", evidence),
        ("predicted_answer", result["predicted"]),
        ("retrieval_path", result["breadcrumb"]),
        ("verdict", result["verdict"]),
    ]:
        print(f"\n--- {field} ---\n{value}")


def _print_summary(counters: Dict[str, int], out_path: Path) -> None:
    n_t, n_f = counters["T"], counters["F"]
    total_judged = n_t + n_f
    print(f"\n{'=' * 60}")
    print(f"Done.  T={n_t}  F={n_f}  ?={counters['?']}  "
          f"skipped={counters['SKIP']}  errored={counters['ERROR']}")
    if total_judged:
        print(f"Accuracy (T / (T+F)) = {n_t}/{total_judged} = {n_t / total_judged:.1%}")
    print(f"Results saved to {out_path}")


def main() -> int:
    args = _parse_args()
    out_path = Path(args.out)
    trees_dir = Path(args.trees_dir)

    try:
        ws, header, col = _load_qa(Path(args.qa_file))
    except ValueError as err:
        print(f"Error: {err}", file=sys.stderr)
        return 1

    out_wb, out_ws = _make_output_workbook(header)
    config = Config()
    routing_llm, answer_llm, judge_llm = _make_llm_clients(config)

    n_total = ws.max_row - 1
    counters: Dict[str, int] = {"T": 0, "F": 0, "?": 0, "SKIP": 0, "ERROR": 0}
    tree_cache: Dict[str, Node] = {}

    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if args.limit and (r - 1) > args.limit:
            break

        result = _process_row(r, n_total, row, col, trees_dir, tree_cache,
                              config, routing_llm, answer_llm, judge_llm, args)
        verdict = result["verdict"]
        counters[verdict] += 1
        out_ws.append(list(row) + [result["predicted"], result["breadcrumb"], verdict])

        if verdict in ("T", "F", "?"):
            _print_progress(r, n_total, row, col, result, counters)
        if verdict == "F" and args.dump_false:
            out_wb.save(out_path)
            _print_f_dump(row, header, col, result)

        if (r - 1) % args.save_every == 0:
            out_wb.save(out_path)

    out_wb.save(out_path)
    _print_summary(counters, out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())